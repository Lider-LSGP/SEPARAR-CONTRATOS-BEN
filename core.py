"""
Núcleo de processamento — sem dependência do Streamlit.

Pode ser usado standalone (CLI/scripts) ou pelo app.py.
"""

from __future__ import annotations

import io
import re
import zipfile
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


MAPEAMENTO_SHEET = "POSTOS"

# Cores da identidade visual LIDER LIMPE
COR_AZUL = "1B3A8C"      # azul escuro do logo
COR_LARANJA = "F47B20"   # laranja do logo
COR_CINZA = "6C7686"     # cinza secundário


# ════════════════════════════════════════════════════════════════════════
# Utilidades
# ════════════════════════════════════════════════════════════════════════
def normalize_text(s) -> str:
    """Normaliza string para comparação (sem acento, upper, trim, sem espaços duplos)."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    s = re.sub(r"\s+", " ", s).upper()
    return s


def normalize_cpf(cpf) -> str:
    """Mantém apenas dígitos do CPF e completa com zeros à esquerda até 11."""
    if cpf is None or (isinstance(cpf, float) and pd.isna(cpf)):
        return ""
    digits = re.sub(r"\D", "", str(cpf))
    if not digits:
        return ""
    return digits.zfill(11)


def safe_filename(name: str) -> str:
    """Remove caracteres inválidos para nome de arquivo Windows/Linux."""
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or "SEM_NOME"


def parse_valor(v):
    """Converte VALOR (float/int/str BR) para float. Vazios viram 0.0."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if not s:
        return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def read_excel_any(file, sheet_name=0, header=None) -> pd.DataFrame:
    """Lê .xls (xlrd) ou .xlsx (openpyxl) de forma robusta."""
    name = getattr(file, "name", str(file)).lower()
    if name.endswith(".xls"):
        engines = ["xlrd", "openpyxl"]
    elif name.endswith(".xlsx"):
        engines = ["openpyxl", "xlrd"]
    else:
        engines = ["xlrd", "openpyxl"]

    last_err = None
    for eng in engines:
        try:
            if hasattr(file, "seek"):
                file.seek(0)
            return pd.read_excel(file, sheet_name=sheet_name, header=header, engine=eng)
        except Exception as e:
            last_err = e
            continue
    raise RuntimeError(f"Não foi possível ler o arquivo Excel. Último erro: {last_err}")


# ════════════════════════════════════════════════════════════════════════
# Núcleo: parsing do Relatório VA e Mapeamento
# ════════════════════════════════════════════════════════════════════════
def carregar_mapeamento(file_or_path) -> dict:
    """
    Lê a aba POSTOS do Mapeamento Sistema e devolve dict:
        { normalize_text(Nome_EasyApp) -> CONTRATO }
    """
    df = read_excel_any(file_or_path, sheet_name=MAPEAMENTO_SHEET, header=0)
    cols_lower = {c.lower().strip(): c for c in df.columns if isinstance(c, str)}
    col_nome = cols_lower.get("nome_easyapp")
    col_contrato = cols_lower.get("contrato")
    if col_nome is None or col_contrato is None:
        raise ValueError(
            "A aba 'POSTOS' do Mapeamento precisa ter as colunas "
            "'Nome_EasyApp' e 'CONTRATO'."
        )

    mp: dict = {}
    for _, row in df.iterrows():
        nome = row[col_nome]
        contrato = row[col_contrato]
        if pd.isna(nome) or pd.isna(contrato):
            continue
        key = normalize_text(nome)
        if not key:
            continue
        if key not in mp:
            mp[key] = str(contrato).strip()
    return mp


def parse_relatorio_va(file):
    """
    Lê a planilha do Relatório VA e retorna (df_limpo, meta).
    df_limpo: NOME | CPF | POSTO | VALOR (linhas-lixo removidas).
    meta: dict com 'periodo', 'codigo_mapa', etc. quando achados.
    """
    raw = read_excel_any(file, sheet_name=0, header=None)

    # 1) Achar linha de cabeçalho com NOME e CPF
    header_row = None
    for i in range(min(15, len(raw))):
        row_vals = [normalize_text(v) for v in raw.iloc[i].tolist()]
        if "NOME" in row_vals and "CPF" in row_vals:
            header_row = i
            break
    if header_row is None:
        header_row = 0

    header_vals = [normalize_text(v) for v in raw.iloc[header_row].tolist()]

    def find_col(*names):
        for nm in names:
            if nm in header_vals:
                return header_vals.index(nm)
        return None

    idx_nome = find_col("NOME")
    idx_cpf = find_col("CPF")
    idx_posto = find_col("POSTO TRABALHO", "POSTO DE TRABALHO", "POSTO")
    idx_lote = find_col("LOTE")
    idx_valor = find_col("VALOR")

    if any(v is None for v in (idx_nome, idx_cpf, idx_posto, idx_valor)):
        raise ValueError(
            "Cabeçalho não reconhecido. Esperado pelo menos: "
            "NOME, CPF, POSTO TRABALHO, VALOR."
        )

    body = raw.iloc[header_row + 1:].copy().reset_index(drop=True)

    # 2) Identificar onde começa o rodapé
    meta = {"periodo": None, "total_geral": None, "codigo_mapa": None}
    footer_idx_candidates = []
    for i in range(len(body)):
        c0 = str(body.iat[i, 0]) if not pd.isna(body.iat[i, 0]) else ""
        c3 = ""
        if idx_lote is not None and idx_lote < body.shape[1]:
            c3 = str(body.iat[i, idx_lote]) if not pd.isna(body.iat[i, idx_lote]) else ""
        low = c0.lower()
        if low.startswith("período") or low.startswith("periodo"):
            meta["periodo"] = c0.split(":", 1)[-1].strip()
            footer_idx_candidates.append(i)
        elif low.startswith("código/mapa") or low.startswith("codigo/mapa"):
            meta["codigo_mapa"] = c0.split(":", 1)[-1].strip()
            footer_idx_candidates.append(i)
        elif low.startswith("empresa:") or low.startswith("relatório de va") or \
             low.startswith("relatorio de va") or low.startswith("data/hora"):
            footer_idx_candidates.append(i)
        elif "total/geral" in c3.lower() or "total geral" in c3.lower():
            footer_idx_candidates.append(i)

    cutoff = min(footer_idx_candidates) if footer_idx_candidates else len(body)

    # 3) Filtrar
    rows = []
    for i in range(cutoff):
        nome_raw = body.iat[i, idx_nome]
        cpf_raw = body.iat[i, idx_cpf]
        posto_raw = body.iat[i, idx_posto]
        valor_raw = body.iat[i, idx_valor]
        lote_raw = body.iat[i, idx_lote] if idx_lote is not None else None

        # separador "Posto de Trabalho:"
        if isinstance(nome_raw, str) and nome_raw.strip().lower().startswith("posto de trabalho"):
            continue
        # subtotal
        if isinstance(lote_raw, str) and "subtotal" in lote_raw.strip().lower():
            continue
        # linha em branco / só obs
        nome_vazio = nome_raw is None or (isinstance(nome_raw, float) and pd.isna(nome_raw))
        cpf_vazio = cpf_raw is None or (isinstance(cpf_raw, float) and pd.isna(cpf_raw))
        if nome_vazio and cpf_vazio:
            continue

        rows.append({
            "NOME": str(nome_raw).strip() if not pd.isna(nome_raw) else "",
            "CPF": normalize_cpf(cpf_raw),
            "POSTO": str(posto_raw).strip() if not pd.isna(posto_raw) else "",
            "VALOR": parse_valor(valor_raw),
        })

    df = pd.DataFrame(rows, columns=["NOME", "CPF", "POSTO", "VALOR"])
    return df, meta


def aplicar_mapeamento(df: pd.DataFrame, mp: dict) -> pd.DataFrame:
    """Adiciona coluna CONTRATO; postos não mapeados → 'SEM_CONTRATO'."""
    def lookup(posto):
        return mp.get(normalize_text(posto), "SEM_CONTRATO")
    df = df.copy()
    df["CONTRATO"] = df["POSTO"].map(lookup)
    return df


# ════════════════════════════════════════════════════════════════════════
# Geração das planilhas por contrato
# ════════════════════════════════════════════════════════════════════════
def build_xlsx_contrato(nome_contrato: str, df_contrato: pd.DataFrame) -> bytes:
    """
    Gera .xlsx em memória com estilo + linha de TOTAL.

    - Contratos normais: NOME | CPF | VALOR        (3 colunas)
    - SEM_CONTRATO:      NOME | CPF | POSTO | VALOR (4 colunas — mostra o
      posto de cada colaborador para facilitar o cadastro no Mapeamento)
    """
    # Define se este arquivo é o "SEM_CONTRATO" (mostra coluna POSTO)
    incluir_posto = (nome_contrato.upper().strip() == "SEM_CONTRATO")

    if incluir_posto:
        headers = ["NOME", "CPF", "POSTO", "VALOR"]
        col_valor = 4
    else:
        headers = ["NOME", "CPF", "VALOR"]
        col_valor = 3
    n_cols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = safe_filename(nome_contrato)[:31] or "Planilha1"
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor=COR_AZUL)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Estilizar cabeçalho
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Linhas de dados
    for _, r in df_contrato.iterrows():
        if incluir_posto:
            ws.append([r["NOME"], r["CPF"], r.get("POSTO", ""), r["VALOR"]])
        else:
            ws.append([r["NOME"], r["CPF"], r["VALOR"]])

    # Formatação das linhas
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=2).number_format = "@"  # CPF como texto
        ws.cell(row=row, column=col_valor).number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
        for col in range(1, n_cols + 1):
            ws.cell(row=row, column=col).border = border
            if col == 1:
                align = "left"
            elif col == 2:
                align = "center"
            elif col == col_valor:
                align = "right"
            else:  # coluna POSTO
                align = "left"
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal=align, vertical="center",
            )

    # Linha de TOTAL
    total_row = last_row + 1
    label_col = col_valor - 1  # célula imediatamente antes do VALOR
    ws.cell(row=total_row, column=label_col, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=label_col).alignment = Alignment(horizontal="right")
    for col in range(1, n_cols + 1):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="F2F2F2")
    total_cell = ws.cell(row=total_row, column=col_valor, value=float(df_contrato["VALOR"].sum()))
    total_cell.font = Font(bold=True, color=COR_AZUL)
    total_cell.number_format = 'R$ #,##0.00'

    # Larguras de coluna
    ws.column_dimensions["A"].width = 45    # NOME
    ws.column_dimensions["B"].width = 16    # CPF
    if incluir_posto:
        ws.column_dimensions["C"].width = 40    # POSTO
        ws.column_dimensions["D"].width = 16    # VALOR
    else:
        ws.column_dimensions["C"].width = 16    # VALOR
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_zip(grupos: dict, mes: int, ano: int) -> bytes:
    """
    grupos: { contrato -> DataFrame }
    Gera ZIP com cada contrato em "CONTRATO      MM-AAAA.xlsx".
    """
    buf = io.BytesIO()
    sufixo = f"{mes:02d}-{ano:04d}"
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for contrato, df_c in sorted(grupos.items()):
            xlsx_bytes = build_xlsx_contrato(contrato, df_c)
            fname = f"{safe_filename(contrato)}      {sufixo}.xlsx"
            zf.writestr(fname, xlsx_bytes)
    buf.seek(0)
    return buf.getvalue()
